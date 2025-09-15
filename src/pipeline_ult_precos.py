import pandas as pd
import requests
from datetime import date, timedelta
from openpyxl import load_workbook

# Caminho dos arquivos
HEADER_CSV = 'data/cabecalho_pedido.csv'
ITEMS_CSV = 'data/item_pedido.csv'
OUTPUT_CSV = "relatorios/relatorio_final.csv"
OUTPUT_XLSX = "relatorios/relatorio_final.xlsx"

BCB_BASE = (
    "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
    "CotacaoMoedaPeriodo(moeda=@moeda,dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
)
BCB_DATE_FORMAT = "%m-%d-%Y"

# Cache de cotações por moeda
rate_cache = {}

# ---------- FUNÇÕES ----------

def ler_csvs(header_path, items_path):
    df_header = pd.read_csv(header_path, dtype=str)
    df_header.columns = [c.lower().strip() for c in df_header.columns]
    df_header['data_pedido'] = pd.to_datetime(df_header['data_pedido'], errors='coerce')

    df_items = pd.read_csv(items_path, dtype=str)
    df_items.columns = [c.lower().strip() for c in df_items.columns]
    df_items['valor_total_item_pedido'] = pd.to_numeric(df_items['valor_total_item_pedido'], errors='coerce')

    # Renomear coluna item_quantidade -> quantidade_item
    if 'item_quantidade' in df_items.columns:
        df_items.rename(columns={'item_quantidade': 'quantidade_item'}, inplace=True)

    return df_header, df_items


def obter_ultima_compra(df_header, df_items):
    df = df_items.merge(df_header, on="codigo_pedido", how="left")
    df_sorted = df.sort_values(['codigo_material', 'data_pedido'], ascending=[True, False])
    df_last = df_sorted.groupby('codigo_material', as_index=False).first()

    df_last = df_last.rename(columns={
        'valor_total_item_pedido': 'ultimo_preco_sem_conversao',
        'moeda': 'moeda_do_pedido',
        'data_pedido': 'data_ultima_compra',
        'codigo_pedido': 'codigo_pedido_referencia'
    })

    return df_last


def consultar_bcb(moeda, data_inicial, data_final):
    """
    Consulta a API do BCB para pegar a última cotação disponível da moeda.
    Se não encontrar para a data solicitada, volta dias anteriores até achar.
    """
    key = f"{moeda}_{data_inicial}_{data_final}"
    if key in rate_cache:
        return rate_cache[key]

    dt_ini = pd.to_datetime(data_inicial)
    dt_fim = pd.to_datetime(data_final)

    dias_tentativa = 10
    for i in range(dias_tentativa):
        data_ini_try = (dt_ini - timedelta(days=i)).strftime(BCB_DATE_FORMAT)
        data_fim_try = (dt_fim - timedelta(days=i)).strftime(BCB_DATE_FORMAT)

        url = f"{BCB_BASE}?@moeda='{moeda}'&@dataInicial='{data_ini_try}'&@dataFinalCotacao='{data_fim_try}'&$format=json"
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            values = resp.json().get('value', [])
            if values:
                values_sorted = sorted(values, key=lambda x: x['dataHoraCotacao'], reverse=True)
                rate_cache[key] = values_sorted[0]
                return values_sorted[0]
        except Exception as e:
            print(f"Erro ao consultar BCB para {moeda}: {e}")

    rate_cache[key] = None
    return None


def converter_para_brl(valor, rate_record):
    if rate_record is None or pd.isna(valor):
        return None
    cotacao = rate_record.get('cotacaoVenda') or rate_record.get('cotacaoCompra')
    if cotacao is None:
        return None
    return round(float(valor) * float(cotacao), 2)


def formatar_moeda(valor, moeda):
    if pd.isna(valor) or valor is None:
        return ""
    if moeda.upper() == 'BRL':
        return f"R${valor:,.2f}"
    elif moeda.upper() == 'USD':
        return f"US${valor:,.2f}"
    elif moeda.upper() == 'EUR':
        return f"€{valor:,.2f}"
    else:
        return f"{valor:,.2f} {moeda}"


def gerar_relatorio(df_last):
    today = date.today()
    valores_brl = []
    datas_cotacao = []

    # Lista de moedas únicas (exceto BRL)
    moedas = df_last['moeda_do_pedido'].str.upper().unique()
    moedas = [m for m in moedas if m != 'BRL']

    # Consulta API apenas uma vez por moeda
    bcb_rates = {}
    for moeda in moedas:
        rate = consultar_bcb(moeda, today.strftime(BCB_DATE_FORMAT), today.strftime(BCB_DATE_FORMAT))
        bcb_rates[moeda] = rate

    # Preenche valores convertidos
    for _, row in df_last.iterrows():
        moeda = str(row['moeda_do_pedido']).upper()
        valor_original_num = row['ultimo_preco_sem_conversao']

        if moeda != 'BRL' and moeda in bcb_rates:
            rate = bcb_rates[moeda]
            if rate:
                valor_brl = converter_para_brl(valor_original_num, rate)
                data_cotacao = pd.to_datetime(rate.get('dataHoraCotacao')).date()
            else:
                valor_brl = None
                data_cotacao = None
        else:
            valor_brl = valor_original_num
            data_cotacao = None

        valores_brl.append(valor_brl)
        datas_cotacao.append(data_cotacao)

    # Atualiza colunas
    df_last['ultimo_preco_brl'] = [formatar_moeda(v, 'BRL') if v is not None else "" for v in valores_brl]
    df_last['ultimo_preco_sem_conversao'] = [
        formatar_moeda(v, m) if v is not None else "" 
        for v, m in zip(df_last['ultimo_preco_sem_conversao'], df_last['moeda_do_pedido'])
    ]
    df_last['data_cotacao_considerada'] = datas_cotacao
    df_last['data_ultima_compra'] = df_last['data_ultima_compra'].dt.date

    # Reorganizar colunas no relatório final
    colunas_final = [
        'codigo_material',
        'ultimo_preco_brl',
        'ultimo_preco_sem_conversao',
        'moeda_do_pedido',
        'data_ultima_compra',
        'codigo_pedido_referencia',
        'data_cotacao_considerada'
    ]
    df_last = df_last[colunas_final]

    # Salvar CSV e Excel
    df_last.to_csv(OUTPUT_CSV, index=False)
    df_last.to_excel(OUTPUT_XLSX, index=False)
    print(f"Relatório gerado: {OUTPUT_CSV} e {OUTPUT_XLSX}")

    # Adicionar formatação no Excel
    adicionar_formato_excel(OUTPUT_XLSX)


def adicionar_formato_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Congelar primeira linha
    ws.freeze_panes = ws['A2']

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)
    print("Excel formatado com filtros e cabeçalho congelado!")


# ---------- EXECUÇÃO ----------

def main():
    df_header, df_items = ler_csvs(HEADER_CSV, ITEMS_CSV)
    df_last = obter_ultima_compra(df_header, df_items)
    gerar_relatorio(df_last)
    print("Pipeline finalizado com sucesso!")


if __name__ == "__main__":
    main()
